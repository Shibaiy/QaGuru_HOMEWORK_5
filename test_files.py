from pypdf import PdfReader
import zipfile
import csv
from openpyxl import load_workbook


def test_PDF():

     with zipfile.ZipFile("resources/archive.zip", "r") as zip_file:
         with zip_file.open('Demo.pdf') as file:
             page = PdfReader(file)
             assert 'Команда UDL' in page.pages[0].extract_text()




def test_CSV():
    with zipfile.ZipFile("resources/archive.zip", "r") as zip_file:
        with zip_file.open('CSV.csv') as file:
            page = file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(page.splitlines()))
            assert ['Timestamp;Value;Quality;Annotation'] == csvreader[0]


def test_XLSX():
    with zipfile.ZipFile("resources/archive.zip", "r") as zip_file:
        with zip_file.open('XLSX.xlsx') as file:
            page = load_workbook(file).active
            assert  page.cell(row=2,column=2).value == 'Dulce'


